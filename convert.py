"""
Azure Pricing Calculator → Cost Estimation Workbook (Production Ready)
"""
import re
import sys
import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ── Style Helpers ────────────────────────────────────────────────────────────
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_SYMBOLS = {"USD": "$", "INR": "₹", "EUR": "€", "GBP": "£", "AUD": "A$"}

def get_num_fmt(currency):
    sym = CURRENCY_SYMBOLS.get(currency, currency)
    return f'"{sym}"#,##0.00'

def _f(bold=False, italic=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)

def _fill(h): return PatternFill("solid", fgColor=h)
def _al(h="left", v="center", w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)

def hdr(c, v, wrap=False):
    c.value=v; c.font=Font(name="Calibri",bold=True,size=10,color="FFFFFF")
    c.fill=_fill("4472C4"); c.alignment=_al("center","center",wrap); c.border=BORDER

def dat(c, v, currency, bold=False, italic=False, align="left", color="000000"):
    c.value=v; c.font=_f(bold,italic,color=color)
    c.alignment=_al(align,"center",True); c.border=BORDER
    if (isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("="))) and align=="right":
        c.number_format = get_num_fmt(currency)

def tot(c, v, currency):
    c.value=v; c.font=_f(bold=True); c.fill=_fill("D9E1F2")
    c.alignment=_al("right","center"); c.border=BORDER
    c.number_format = get_num_fmt(currency)

def widths(ws, d):
    for col, w in d.items(): ws.column_dimensions[col].width = w

SVC_MAP = {
    "virtual machines": "Virtual Machines",
    "managed disks": "Managed Disks",
    "azure backup": "Azure Backup",
    "load balancer": "Load Balancer",
    "application gateway": "Application Gateway",
    "azure firewall": "Azure Firewall",
    "vpn gateway": "VPN Gateway",
    "storage": "Storage Accounts",
    "storage accounts": "Storage Accounts",
    "sql database": "SQL",
    "azure sql": "SQL",
    "ip addresses": "Public IP Addresses",
    "bandwidth": "Bandwidth",
    "azure monitor": "Azure Monitor",
    "key vault": "Key Vault",
}

SHEET_ORDER = [
    "Virtual Machines", "Managed Disks", "Public IP Addresses",
    "Load Balancer", "Application Gateway", "Azure Firewall", "VPN Gateway",
    "Storage Accounts", "Azure Backup", "SQL", "Bandwidth",
    "Azure Monitor", "Key Vault", "Others"
]

SKIP = {"support", "disclaimer", "total", "licensing program", "billing account", "billing profile"}

def arm_region(display):
    val = str(display).lower().strip()
    return val.replace(" ", "")

# ── API Setup & Helpers ──────────────────────────────────────────────────
API = "https://prices.azure.com/api/retail/prices"

def get_http_session():
    session = requests.Session()
    retries = Retry(total=5, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504])
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session

def _api(session, cache, filt, currency="INR"):
    key = filt + currency
    if key in cache: return cache[key]
    try:
        r = session.get(API, params={"api-version":"2023-01-01-preview", "$filter":filt, "currencyCode":currency}, timeout=15)
        r.raise_for_status()
        items = r.json().get("Items", [])
        cache[key] = items
        return items
    except Exception as e:
        log.warning(f"API Fetch Error: {e}")
        return []

def _hourly_to_monthly(price): return price * 730

def get_exact_license_name(desc):
    desc_l = desc.lower()
    parts = re.split(r'[,;]', desc)
    sql_name = os_name = None
    
    is_ahb = any(kw in desc_l for kw in ["hybrid benefit", "ahb", "bring your own license", "byol"])
    
    prem_kws = ["red hat", "rhel", "suse", "sles", "ubuntu pro", "ubuntu advantage"]
    for p in parts:
        p_lower = p.lower()
        if "sql" in p_lower:
            sql_name = re.sub(r'\s*\([^)]*\)', '', p).strip()
        if any(kw in p_lower for kw in prem_kws) and "sql" not in p_lower:
            os_name = re.sub(r'\s*\([^)]*\)', '', p).strip()
            if os_name.lower().startswith("linux "):
                os_name = os_name[6:].strip()
                
    os_type = "Windows" if "windows" in desc_l else "Linux"
    return sql_name, os_name, os_type, is_ahb

def get_safe_sku_fallbacks(sku):
    fallbacks = []
    f1 = re.sub(r'(?i)s(_v\d+)$', r'\1', sku)
    if f1 != sku: fallbacks.append(f1)
    f2 = re.sub(r'(?i)ds(\d+_v\d+)$', r'd\1', sku)
    if f2 != sku: fallbacks.append(f2)
    return fallbacks

def get_vm_pricing(session, cache, sku, region_display, is_spot, currency="INR"):
    region = arm_region(region_display)
    
    def fetch(s, ptype):
        filt = f"armSkuName eq '{s}' and armRegionName eq '{region}' and priceType eq '{ptype}' and serviceName eq 'Virtual Machines'"
        return _api(session, cache, filt, currency)

    def get_items(s, ptype):
        items = fetch(s, ptype)
        if items: return items
        for fallback in get_safe_sku_fallbacks(s):
            items = fetch(fallback, ptype)
            if items: return items
        return []

    payg_items = get_items(sku, "Consumption")
    ri_items = get_items(sku, "Reservation")
    
    if not payg_items:
        base_sku = re.sub(r'-\d+', '', sku)
        if base_sku != sku:
            payg_items = get_items(base_sku, "Consumption")
            ri_items = get_items(base_sku, "Reservation")

    def _get_price(items, must_be_win=False):
        cands = []
        for i in items:
            prod, meter = i.get("productName", "").lower(), i.get("meterName", "").lower()
            if "low priority" in meter: continue
            if is_spot and "spot" not in meter: continue
            if not is_spot and "spot" in meter: continue
            
            is_win_prod = "windows" in prod
            if must_be_win and not is_win_prod: continue
            if not must_be_win and is_win_prod: continue
            if i.get("retailPrice", 0) > 0: cands.append(i["retailPrice"])
        return min(cands) if cands else None

    linux_hr = _get_price(payg_items, must_be_win=False)
    win_hr   = _get_price(payg_items, must_be_win=True)

    result = {
        "compute_payg": _hourly_to_monthly(linux_hr) if linux_hr else None,
        "windows_tot": _hourly_to_monthly(win_hr) if win_hr else None,
        "compute_ri1": None,
        "compute_ri3": None
    }

    if not is_spot:
        ri1_cands = [i for i in ri_items if i.get("reservationTerm") == "1 Year"]
        ri1_val = _get_price(ri1_cands, must_be_win=False)
        if ri1_val is not None: result["compute_ri1"] = ri1_val / 12

        ri3_cands = [i for i in ri_items if i.get("reservationTerm") == "3 Years"]
        ri3_val = _get_price(ri3_cands, must_be_win=False)
        if ri3_val is not None: result["compute_ri3"] = ri3_val / 36

    return result

def extract_vm_sku(desc):
    if not desc: return None
    m = re.match(r'^\s*[\d,]+\s+([^()]+)\(', desc)
    if m:
        norm = re.sub(r'\s+', '_', m.group(1).strip())
        return norm if norm.lower().startswith("standard_") else "Standard_" + norm
    return None

def extract_quantity(desc):
    if not desc: return 1
    m = re.match(r'^\s*([0-9,]+)\s+', desc)
    return max(1, int(m.group(1).replace(',', ''))) if m else 1

# ── Parsing ──────────────────────────────────────────────────────────────────
def parse_all_formats(wb):
    rows = []
    for sname in wb.sheetnames:
        if sname.lower() == "summary": continue
        ws = wb[sname]
        in_data = False
        
        def _flt(row_data, idx, fallback):
            if len(row_data) > idx and isinstance(row_data[idx], (int, float)) and row_data[idx] > 0:
                return float(row_data[idx])
            return float(fallback)

        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None: continue
            
            svc_cat_raw = str(r[0]).strip().lower()
            if svc_cat_raw == "service category":
                in_data = True
                continue
                
            if not in_data: continue
            
            svc_cat, svc_type, region, desc = str(r[0] or "").strip(), str(r[1] or "").strip(), str(r[3] or "").strip(), str(r[4] or "").strip()
            cost_raw = r[5]
            
            if not desc or desc.lower() == "total" or svc_cat.lower() in SKIP or region.lower() in SKIP or not isinstance(cost_raw, (int, float)):
                continue
                
            rows.append({
                "svc_cat": svc_cat, "svc_type": svc_type, "cust_name": str(r[2] or "").strip(),
                "region": region, "desc": desc, "payg": float(cost_raw),
                "ri1": _flt(r, 6, cost_raw), "ri3": _flt(r, 7, cost_raw), 
                "remarks": "", "sub_rows": []
            })
    return rows

def classify(rows):
    buckets = {}
    for r in rows:
        key = str(r.get("svc_type") or "").lower().strip()
        sheet = SVC_MAP.get(key, "Others")
        if sheet == "Others":
            for k, v in SVC_MAP.items():
                if k in key: sheet = v; break
        buckets.setdefault(sheet, []).append(r)
    return buckets

def enrich_vms_concurrent(vm_rows, currency="INR"):
    log.info(f"Querying Azure Retail Pricing API concurrently for {len(vm_rows)} VMs in {currency}...")
    session = get_http_session()
    cache = {}

    locked_fx = 1.0
    if currency != "USD":
        for row in vm_rows:
            desc = row["desc"]
            sku = extract_vm_sku(desc)
            sql_lbl, _, _, is_ahb = get_exact_license_name(desc)
            prem_kws = ["red hat", "rhel", "suse", "sles", "ubuntu pro", "ubuntu advantage"]
            has_prem = any(kw in desc.lower() for kw in prem_kws)
            
            if sku and not has_prem and not sql_lbl:
                qty = extract_quantity(desc)
                p_usd = get_vm_pricing(session, cache, sku, row["region"], "spot" in desc.lower(), "USD")
                usd_comp = (p_usd.get("compute_payg") or 0) * qty
                usd_win = (p_usd.get("windows_tot") or 0) * qty if detect_os(desc) == "Windows" and not is_ahb else 0
                usd_tot = usd_comp + usd_win
                orig_payg = row.get("payg", 0)
                
                if usd_tot > 0 and orig_payg > 0:
                    locked_fx = orig_payg / usd_tot
                    break

    def process_row(row):
        try: 
            desc, region = row["desc"], row["region"]
            qty = extract_quantity(desc)
            sku = extract_vm_sku(desc)

            sql_exact, os_exact, os_type, is_ahb = get_exact_license_name(desc)
            has_premium_os = bool(os_exact)
            has_sql = bool(sql_exact)
            is_spot = "spot" in desc.lower()
            
            row["sql_lbl_exact"] = sql_exact or "SQL License"
            row["os_lbl_exact"]  = os_exact or f"Premium OS License"
            row["api"] = {}
            remarks = []

            if is_spot: remarks.append("Spot VM (RIs N/A)")
            if is_ahb: remarks.append("AHB/BYOL Applied")

            if not sku:
                remarks.append("SKU Parsing Failed")
                if has_premium_os or has_sql: row["api"] = {"is_standalone": True}
                row["remarks"] = " | ".join(remarks)
                return row

            p_usd = get_vm_pricing(session, cache, sku, region, is_spot, "USD")
            usd_comp = (p_usd.get("compute_payg") or 0) * qty
            usd_win  = (p_usd.get("windows_tot") or 0) * qty
            usd_ri1  = (p_usd.get("compute_ri1") or 0) * qty
            usd_ri3  = (p_usd.get("compute_ri3") or 0) * qty

            orig_payg = row.get("payg", 0)
            orig_ri1, orig_ri3 = row.get("ri1", orig_payg), row.get("ri3", orig_payg)

            if usd_comp == 0:
                remarks.append("API Lookup Failed - Using Raw Values")
                row["api"] = {
                    "compute_payg_final": orig_payg, "win_lic_payg_final": 0, 
                    "prem_os_payg_final": 0, "sql_payg_final": 0,
                    "compute_ri1": orig_ri1, "compute_ri3": orig_ri3
                }
                row["remarks"] = " | ".join(remarks)
                return row

            comp_payg, win_payg, prem_os_payg, sql_payg = orig_payg, 0, 0, 0

            # --- PROPORTIONAL FX SPLIT ENGINE ---
            if os_type == "Windows" and not is_ahb and usd_win > 0:
                win_ratio = usd_win / (usd_comp + usd_win)
                win_payg = orig_payg * win_ratio
                comp_payg = orig_payg - win_payg
            elif has_premium_os:
                comp_payg = usd_comp * locked_fx
                prem_os_payg = max(0, orig_payg - comp_payg)
                # Visual Merge
                comp_payg += prem_os_payg
                prem_os_payg = 0 
            
            if has_sql:
                if orig_payg > (comp_payg + win_payg + prem_os_payg):
                    sql_payg = orig_payg - (comp_payg + win_payg + prem_os_payg)
                else:
                    sql_payg = orig_payg * 0.40
                    comp_payg = orig_payg * 0.60

            license_total = win_payg + prem_os_payg + sql_payg
            
            # --- FLAWLESS RI COMPUTATION ---
            if is_spot:
                comp_ri1 = comp_payg
                comp_ri3 = comp_payg
            else:
                if usd_ri1 > 0:
                    comp_ri1 = (usd_ri1 * locked_fx) + prem_os_payg
                else:
                    comp_ri1 = max(0, orig_ri1 - license_total)
                    
                if usd_ri3 > 0:
                    comp_ri3 = (usd_ri3 * locked_fx) + prem_os_payg
                else:
                    comp_ri3 = max(0, orig_ri3 - license_total)

            row["api"] = {
                "compute_payg_final": comp_payg,
                "win_lic_payg_final": win_payg,
                "prem_os_payg_final": prem_os_payg,
                "sql_payg_final": sql_payg,
                "compute_ri1": comp_ri1,
                "compute_ri3": comp_ri3
            }
            row["remarks"] = " | ".join(remarks)
            return row

        except Exception as e:
            log.error(f"Worker Exception on row: {e}")
            row["api"] = {"error": True}
            row["remarks"] = f"Processing Error: {str(e)[:50]}"
            return row

    with ThreadPoolExecutor(max_workers=min(20, len(vm_rows))) as executor:
        futures = {executor.submit(process_row, r): r for r in vm_rows}
        for future in as_completed(futures):
            try: future.result() 
            except Exception as e: log.error(f"Pool Exception: {e}")

    return vm_rows

# ── Output ───────────────────────────────────────────────────────────────────
def write_res_header(ws):
    ws.merge_cells("F1:H1")
    hdr(ws["F1"], "Monthly Cost")
    for addr in ["G1","H1"]: ws[addr].border=BORDER
    for ci, h in enumerate(["Service category","Service type","Custom name", "Region","Description","PAYG", "1 Year RI Model","3 Year RI Model","Remarks"], 1):
        hdr(ws.cell(2, ci), h, wrap=True)
    ws.row_dimensions[2].height = 28.8
    ws.freeze_panes = "A3"

def write_vm_sheet(wb, rows, currency):
    ws = wb.create_sheet("Virtual Machines")
    write_res_header(ws)
    widths(ws, {"A":15, "B":15, "C":14, "D":12, "E":55, "F":15, "G":15, "H":15, "I":42})
    ri = 3

    for row in rows:
        p = row.get("api", {})
        if p.get("is_standalone") or p.get("error"):
            payg, ri1, ri3 = row.get("payg", 0), row.get("ri1", 0), row.get("ri3", 0)
            vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], payg, ri1, ri3, row.get("remarks","")]
            for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, currency, align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1
            continue

        comp_payg = p.get("compute_payg_final", row.get("payg",0))
        win_payg  = p.get("win_lic_payg_final", 0)
        prem_os_payg = p.get("prem_os_payg_final", 0)
        sql_payg  = p.get("sql_payg_final", 0)
        
        comp_ri1  = p.get("compute_ri1", comp_payg)
        comp_ri3  = p.get("compute_ri3", comp_payg)

        vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], comp_payg, comp_ri1, comp_ri3, row.get("remarks","")]
        for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, currency, align="right" if ci>=6 and isinstance(v,float) else "left")
        ri += 1

        if win_payg > 0:
            sub = ["","","","","Windows License", win_payg, win_payg, win_payg, "License Cost (Not discounted)"]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, currency, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        if prem_os_payg > 0:
            sub = ["","","","", row.get("os_lbl_exact", "Premium OS License"), prem_os_payg, prem_os_payg, prem_os_payg, "License Cost (Not discounted)"]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, currency, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

        if sql_payg > 0:
            sub = ["","","","", row.get("sql_lbl_exact", "SQL License"), sql_payg, sql_payg, sql_payg, "License Cost (Not discounted)"]
            for ci, v in enumerate(sub, 1): dat(ws.cell(ri, ci), v, currency, italic=True, color="595959", align="right" if ci>=6 and isinstance(v,float) else "left")
            ri += 1

    ws.cell(ri, 5, "Total").font = _f(bold=True)
    ws.cell(ri, 5).border = BORDER
    for ci, col_let in [(6, 'F'), (7, 'G'), (8, 'H')]:
        tot(ws.cell(ri, ci), f"=SUM({col_let}3:{col_let}{ri-1})", currency)
    return ri

def write_generic_sheet(wb, sheet_name, rows, currency):
    ws = wb.create_sheet(sheet_name)
    write_res_header(ws)
    widths(ws, {"A":15, "B":14, "C":22, "D":12, "E":60, "F":15, "G":15, "H":15, "I":40})

    ri = 3
    for row in rows:
        payg = row.get("payg", 0)
        ri1, ri3 = row.get("ri1", payg), row.get("ri3", payg)
        vals = [row["svc_cat"], row["svc_type"], row["cust_name"], row["region"], row["desc"], payg, ri1, ri3, row.get("remarks","")]
        for ci, v in enumerate(vals, 1): dat(ws.cell(ri, ci), v, currency, align="right" if ci>=6 and isinstance(v,float) else "left")
        ri += 1

    ws.cell(ri, 5, "Total").font = _f(bold=True)
    ws.cell(ri, 5).border = BORDER
    for ci, col_let in [(6, 'F'), (7, 'G'), (8, 'H')]:
        tot(ws.cell(ri, ci), f"=SUM({col_let}3:{col_let}{ri-1})", currency)
    return ri

def write_summary(wb, totals, currency):
    ws = wb.create_sheet("Summary", 0)
    ws.merge_cells("A1:A2"); ws.merge_cells("B1:B2"); ws.merge_cells("C1:E1"); ws.merge_cells("F1:F2")
    
    for addr, val, al in [("A1","Sl No","left"),("B1","Service Name","left"), ("C1","Monthly Cost","center"),("F1","Remarks","left")]:
        c=ws[addr]; c.value=val; c.font=_f(bold=True); c.alignment=_al(al,"center"); c.border=BORDER
    for addr, val in [("C2","PAYG"),("D2","1 YR RI Model"),("E2","3 YR RI Model")]:
        c=ws[addr]; c.value=val; c.font=_f(bold=True); c.alignment=_al("center","center"); c.border=BORDER

    ri = 3
    for sl, (sname, tot_row) in enumerate(totals.items(), 1):
        c=ws.cell(ri,1,sl); c.border=BORDER; c.alignment=_al("center")
        c=ws.cell(ri,2,sname); c.border=BORDER
        
        dat(ws.cell(ri, 3), f"='{sname}'!F{tot_row}", currency, align="right")
        dat(ws.cell(ri, 4), f"='{sname}'!G{tot_row}", currency, align="right")
        dat(ws.cell(ri, 5), f"='{sname}'!H{tot_row}", currency, align="right")
        ws.cell(ri,6).border=BORDER
        ri += 1

    ws.cell(ri, 2, "Total").font = _f(bold=True)
    ws.cell(ri, 2).border = BORDER
    ws.cell(ri, 1).border = BORDER
    ws.cell(ri, 6).border = BORDER
    
    tot(ws.cell(ri, 3), f"=SUM(C3:C{ri-1})", currency)
    tot(ws.cell(ri, 4), f"=SUM(D3:D{ri-1})", currency)
    tot(ws.cell(ri, 5), f"=SUM(E3:E{ri-1})", currency)

    widths(ws, {"A":5.5, "B":22, "C":14, "D":14, "E":14, "F":48})

def convert(input_path, output_path, currency="INR"):
    wb_in = load_workbook(input_path, data_only=True)
    rows = parse_all_formats(wb_in)
    
    if not rows:
        raise ValueError("No data rows found. Ensure this is an unmodified Azure Pricing Calculator export.")

    buckets = classify(rows)
    if "Virtual Machines" in buckets:
        enrich_vms_concurrent(buckets["Virtual Machines"], currency)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    totals = {}

    for sname in SHEET_ORDER:
        if sname not in buckets: continue
        if sname == "Virtual Machines":
            totals[sname] = write_vm_sheet(wb_out, buckets[sname], currency)
        else:
            totals[sname] = write_generic_sheet(wb_out, sname, buckets[sname], currency)

    write_summary(wb_out, totals, currency)
    wb_out.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert.py input.xlsx [output.xlsx]")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "output.xlsx")
